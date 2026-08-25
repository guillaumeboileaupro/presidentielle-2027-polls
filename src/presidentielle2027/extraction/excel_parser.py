from __future__ import annotations

import re
from datetime import date
from itertools import zip_longest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from presidentielle2027.extraction.canonicalization import (
    canonicalize_candidate_fields,
    is_generic_bloc_label,
)

PARSING_DIAGNOSTIC_COLUMNS = [
    "fieldwork_date_raw",
    "parse_status",
    "parse_error",
    "estimate_percent_original",
    "estimate_percent_corrected",
    "percentage_correction_applied",
    "percentage_correction_factor",
    "percentage_correction_reason",
    "scenario_total_before",
    "scenario_total_after",
]


def load_workbook_sheets(workbook_path: Path) -> dict[str, pd.DataFrame]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheets: dict[str, pd.DataFrame] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            sheets[sheet_name] = pd.DataFrame()
            continue
        header = [str(cell) if cell is not None else "" for cell in values[0]]
        rows = values[1:]
        sheets[sheet_name] = pd.DataFrame(rows, columns=header)
    return sheets


def _parse_fieldwork_dates(value: str, fallback_year: int | None = None) -> tuple[str | None, str | None]:
    if not isinstance(value, str) or not value.strip():
        return None, None
    normalized_value = value.replace("–", "-").replace("—", "-").strip()
    month_map = {
        "janvier": 1, "january": 1,
        "fevrier": 2, "février": 2, "february": 2, "feb": 2,
        "mars": 3, "march": 3, "mar": 3,
        "avril": 4, "april": 4, "apr": 4,
        "mai": 5, "may": 5,
        "juin": 6, "june": 6, "jun": 6,
        "juillet": 7, "july": 7, "jul": 7,
        "aout": 8, "août": 8, "august": 8, "aug": 8,
        "septembre": 9, "september": 9, "sept": 9, "sep": 9,
        "octobre": 10, "october": 10, "oct": 10,
        "novembre": 11, "november": 11, "nov": 11,
        "decembre": 12, "décembre": 12, "december": 12, "dec": 12,
    }
    year_match = re.search(r"(20\d{2})", normalized_value)
    inferred_year = int(year_match.group(1)) if year_match else fallback_year

    match = re.search(r"(\d{1,2})-(\d{1,2})\s+([A-Za-zéûôîàç]+)(?:\s+(20\d{2}))?", normalized_value, flags=re.IGNORECASE)
    if match:
        start_day, end_day, month_name, year = match.groups()
        month = month_map.get(month_name.lower())
        resolved_year = int(year) if year is not None else inferred_year
        if month is not None and resolved_year is not None:
            start = pd.Timestamp(year=int(resolved_year), month=month, day=int(start_day)).date().isoformat()
            end = pd.Timestamp(year=int(resolved_year), month=month, day=int(end_day)).date().isoformat()
            return start, end
    match = re.search(r"(\d{1,2})\s+([A-Za-zéûôîàç]+)(?:\s+(20\d{2}))?", normalized_value, flags=re.IGNORECASE)
    if match:
        day, month_name, year = match.groups()
        month = month_map.get(month_name.lower())
        resolved_year = int(year) if year is not None else inferred_year
        if month is not None and resolved_year is not None:
            iso = pd.Timestamp(year=int(resolved_year), month=month, day=int(day)).date().isoformat()
            return iso, iso
    parsed = pd.to_datetime(normalized_value, errors="coerce", dayfirst=True)
    if pd.notna(parsed):
        iso = parsed.date().isoformat()
        return iso, iso

    return None, None


def _extract_publication_date_from_poll_id(poll_id: str) -> str | None:
    match = re.search(r"(20\d{2})-(\d{2})-(\d{2})", poll_id)
    if not match:
        return None
    year, month, day = match.groups()
    return f"{year}-{month}-{day}"


def _publication_year_from_poll_id(poll_id: object) -> int | None:
    publication_date = _extract_publication_date_from_poll_id(str(poll_id or ""))
    if not publication_date:
        return None
    return int(publication_date[:4])


def _normalize_company_name(name: str) -> str:
    mapping = {
        "Harris": "Harris Interactive",
        "Ifop": "Ifop",
        "Elabe": "Elabe",
        "Odoxa": "Odoxa",
        "Cluster17": "Cluster17",
    }
    return mapping.get(str(name).strip(), str(name).strip())


def _parse_sample_size(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "/" in text:
        text = text.split("/", 1)[0]
    digits = re.sub(r"[^\d]", "", text)
    return int(digits) if digits else None


def _guess_party_and_family(label: str) -> tuple[str | None, str | None, str]:
    text = label.strip()
    if "=" in text:
        party, candidate = [chunk.strip() for chunk in text.split("=", 1)]
        candidate, party, family = canonicalize_candidate_fields(candidate, party, party)
        return party, family, candidate
    family_map = {
        "Arthaud": "far_left",
        "Poutou": "far_left",
        "Roussel": "left",
        "Mélenchon": "left",
        "Melenchon": "left",
        "Tondelier": "greens",
        "Glucksmann": "centre_left",
        "Philippe": "centre",
        "Attal": "centre",
        "Villepin": "gaullist_right",
        "Retailleau": "right",
        "Dupont-Aignan": "sovereigntist_right",
        "Bardella": "nationalist_right",
        "Le Pen": "nationalist_right",
        "Zemmour": "far_right",
        "Autres": "other",
    }
    candidate, party, family = canonicalize_candidate_fields(text, None, family_map.get(text))
    return party, family, candidate


def _tokenize_score_vector(raw_vector: str) -> list[str]:
    return [token.strip() for token in str(raw_vector).replace("–", " - ").replace("—", " - ").split() if token.strip()]


def _clean_candidate_name(name: str) -> str:
    return (
        str(name)
        .replace("Édouard", "Edouard")
        .replace("Éric", "Eric")
        .replace("É", "E")
        .strip()
    )


def _strip_wikipedia_annotations(text: object) -> str:
    return re.sub(r"\[[^\]]+\]", "", str(text or "")).strip()


def _extract_candidate_and_party_from_label(label: str, header_hint: str | None = None) -> tuple[str, str | None, str | None]:
    cleaned = _clean_candidate_name(_strip_wikipedia_annotations(label))
    party_hint = None
    match = re.match(r"^(.*?)\s*\(([^)]+)\)$", cleaned)
    if match:
        cleaned = match.group(1).strip()
        party_hint = match.group(2).strip()
    if header_hint:
        header_hint = header_hint.strip()
        if header_hint.startswith("Candidat "):
            party_hint = header_hint.replace("Candidat ", "").strip()
    candidate_name, candidate_party, political_family = canonicalize_candidate_fields(cleaned, party_hint, None)
    return candidate_name, candidate_party, political_family


def _parse_raw_poll_percent(value: object) -> float | None:
    text = str(value or "").replace("\xa0", " ").strip()
    if not text or text in {"—", "-", "nan", "NaN"}:
        return None
    if text.startswith("<"):
        return 0.5
    match = re.search(r"(\d+(?:[.,]\d+)?)", text)
    if not match:
        return None
    try:
        return float(match.group(1).replace(",", "."))
    except ValueError:
        return None

def _poll_percentage_options(raw_text: object, parsed_value: float) -> list[float]:
    """Return plausible percentages when an HTML decimal separator was lost."""
    text = str(raw_text or "").replace("\xa0", " ").strip()
    numeric_match = re.search(r"(\d+(?:[.,]\d+)?)", text)
    token = numeric_match.group(1) if numeric_match else ""
    if text.startswith("<"):
        return [0.5]
    if "," in token or ("." in token and not token.endswith(".0")):
        return [float(parsed_value)]

    value = float(parsed_value)
    if len(token) > 1 and token.startswith("0"):
        return [value / 10.0]

    original_value = value
    while value > 100.0:
        value /= 10.0
    options = [value]
    decimal_artifact = token.endswith(".0")
    if value >= 10.0 or decimal_artifact:
        options.append(value / 10.0)
    if original_value > 100.0 or decimal_artifact:
        options.append(value / 100.0)
    return list(dict.fromkeys(round(option, 4) for option in options))


def _percentage_plausibility_penalty(party: object, value: float) -> float:
    party_code = str(party or "").strip()
    plausible_ranges = {
        "LO": (0.0, 5.0),
        "NPA-A": (0.0, 5.0),
        "LFI": (5.0, 30.0),
        "PCF": (0.0, 12.0),
        "EELV": (0.0, 15.0),
        "PS": (2.0, 25.0),
        "PP": (2.0, 25.0),
        "RE": (2.0, 35.0),
        "HOR": (2.0, 35.0),
        "ENS": (2.0, 35.0),
        "LR": (2.0, 25.0),
        "DLF": (0.0, 10.0),
        "RN": (15.0, 50.0),
        "REC": (0.0, 15.0),
    }
    minimum, maximum = plausible_ranges.get(party_code, (0.0, 50.0))
    if value < minimum:
        return minimum - value
    if value > maximum:
        return value - maximum
    return 0.0


def _reconstruct_scenario_percentages(group: pd.DataFrame) -> tuple[pd.Series, bool]:
    """Choose decimal variants whose scenario sum is closest to 100."""
    def state_rank(state: tuple[float, float, list[float]]) -> tuple[float, float, float]:
        distance = abs(state[0] - 100.0)
        total_penalty = 0.0 if distance <= 1.0 else distance
        return total_penalty, state[1], distance

    states: list[tuple[float, float, list[float]]] = [(0.0, 0.0, [])]
    for _, row in group.iterrows():
        parsed_value = float(row["estimate_percent"])
        options = _poll_percentage_options(row.get("raw_text_context"), parsed_value)
        candidates: dict[float, tuple[float, float, list[float]]] = {}
        for running_total, running_cost, selected in states:
            for option_index, option in enumerate(options):
                total = round(running_total + option, 4)
                if total > 130.0:
                    continue
                plausibility = _percentage_plausibility_penalty(row.get("candidate_party"), option)
                candidate = (
                    total,
                    running_cost + option_index + (plausibility * 10.0),
                    [*selected, option],
                )
                previous = candidates.get(total)
                if previous is None or candidate[1] < previous[1]:
                    candidates[total] = candidate
        states = sorted(
            candidates.values(),
            key=lambda state: (state[1], abs(state[0] - 100.0)),
        )[:2048]
    if not states:
        return pd.to_numeric(group["estimate_percent"], errors="coerce"), True
    best = min(states, key=state_rank)
    best_rank = state_rank(best)
    tied = [
        state
        for state in states
        if state_rank(state) == best_rank and state[2] != best[2]
    ]
    return pd.Series(best[2], index=group.index, dtype=float), bool(tied)


def _correct_poll_units_by_scenario(frame: pd.DataFrame) -> pd.DataFrame:
    """Reconstruct percentages at scenario level without discarding polls.

    Wikipedia's HTML tables occasionally lose decimal separators during CSV
    extraction (10.5 becomes 105) and repeat cells created with ``colspan``.
    Exact technical duplicates are retained and marked, then a bounded search
    chooses plausible decimal variants whose total is closest to 100 percent.
    """
    if frame.empty:
        return frame.copy()
    corrected = frame.copy()
    numeric = pd.to_numeric(corrected["estimate_percent"], errors="coerce")
    corrected.loc[numeric.notna(), "estimate_percent"] = numeric.loc[numeric.notna()]
    corrected["estimate_percent_original"] = numeric
    corrected["estimate_percent_corrected"] = numeric
    corrected["percentage_correction_applied"] = False
    corrected["percentage_correction_factor"] = 1.0
    corrected["percentage_correction_reason"] = "unchanged"
    corrected["scenario_total_before"] = float("nan")
    corrected["scenario_total_after"] = float("nan")
    if "parse_status" not in corrected:
        corrected["parse_status"] = "parsed"
    if "parse_error" not in corrected:
        corrected["parse_error"] = None
    scenario_columns = ["poll_id", "round", "scenario_name"]
    raw_context = corrected["raw_text_context"].fillna("").astype(str)
    annotated_context = raw_context.str.contains(r"\[[^\]]+\]", regex=True)
    repeated_merged_cell = corrected.duplicated(
        subset=[*scenario_columns, "raw_text_context"],
        keep="first",
    )
    exact_technical_duplicate = corrected.duplicated(
        subset=[*scenario_columns, "candidate_name", "raw_text_context"],
        keep="first",
    )
    merged_cell_rows = corrected.loc[annotated_context].groupby(
        [*scenario_columns, "raw_text_context"],
        dropna=False,
    )
    for _, merged_group in merged_cell_rows:
        if len(merged_group.index) <= 1:
            continue
        first_index = merged_group.index[0]
        corrected.at[first_index, "candidate_name"] = "NFP"
        corrected.at[first_index, "candidate_party"] = "NFP"
        corrected.at[first_index, "political_family"] = "generic_bloc"
    technical_duplicates = exact_technical_duplicate | (annotated_context & repeated_merged_cell)
    corrected.loc[technical_duplicates, "estimate_percent"] = None
    corrected.loc[technical_duplicates, "estimate_percent_corrected"] = None
    corrected.loc[technical_duplicates, "parse_status"] = "technical_duplicate"
    corrected.loc[technical_duplicates, "parse_error"] = "repeated HTML colspan cell retained for traceability"
    corrected.loc[technical_duplicates, "percentage_correction_reason"] = "technical_duplicate_excluded"

    for _, indexes in corrected.groupby(scenario_columns, dropna=False).groups.items():
        group = corrected.loc[list(indexes)]
        usable = group.loc[
            ~group["parse_status"].isin({"technical_duplicate", "not_tested"})
        ].copy()
        values_before = pd.to_numeric(usable["estimate_percent"], errors="coerce")
        total_before = float(values_before.sum(min_count=1))
        corrected.loc[group.index, "scenario_total_before"] = total_before
        if usable.empty:
            corrected.loc[group.index, "scenario_total_after"] = total_before
            continue
        if values_before.isna().any():
            corrected.loc[group.index, "percentage_correction_reason"] = "ambiguous_missing_value"
            corrected.loc[group.index, "scenario_total_after"] = total_before
            continue
        if total_before <= 101.0 and bool(values_before.le(100.0).all()):
            corrected.loc[group.index, "scenario_total_after"] = total_before
            continue

        reconstructed, ambiguous = _reconstruct_scenario_percentages(usable)
        option_signatures = usable.apply(
            lambda row: (
                tuple(_poll_percentage_options(row.get("raw_text_context"), float(row["estimate_percent"]))),
                str(row.get("candidate_party") or ""),
            ),
            axis=1,
        )
        for _, equivalent_indexes in option_signatures.groupby(option_signatures).groups.items():
            equivalent_indexes = list(equivalent_indexes)
            if len(equivalent_indexes) > 1 and reconstructed.loc[equivalent_indexes].nunique() > 1:
                ambiguous = True
        if ambiguous:
            corrected.loc[group.index, "percentage_correction_reason"] = "ambiguous_multiple_solutions"
            corrected.loc[group.index, "scenario_total_after"] = total_before
            continue

        corrected.loc[reconstructed.index, "estimate_percent"] = reconstructed
        corrected.loc[reconstructed.index, "estimate_percent_corrected"] = reconstructed
        originals = pd.to_numeric(
            corrected.loc[reconstructed.index, "estimate_percent_original"],
            errors="coerce",
        )
        changed = ~originals.eq(reconstructed)
        changed_indexes = reconstructed.index[changed]
        corrected.loc[changed_indexes, "percentage_correction_applied"] = True
        corrected.loc[changed_indexes, "percentage_correction_factor"] = (
            reconstructed.loc[changed_indexes] / originals.loc[changed_indexes]
        )
        corrected.loc[group.index, "percentage_correction_reason"] = "scenario_total_aberrant"
        corrected.loc[group.index, "scenario_total_after"] = float(reconstructed.sum())

    assert len(corrected) == len(frame)
    assert corrected.index.equals(frame.index)
    return corrected


def _row_looks_like_poll(pollster: object, date_text: object, sample_size: object) -> bool:
    if _parse_sample_size(sample_size) is None:
        return False
    pollster_text = str(pollster or "").strip()
    date_label = str(date_text or "").strip()
    if not pollster_text or not date_label:
        return False
    if len(pollster_text) > 40 and " " in pollster_text:
        return False
    return True


def _latest_wikipedia_fr_2027_table_files(raw_dir: Path) -> list[Path]:
    pattern = re.compile(r"wikipedia-fr-2027-polls-(\d{8}T\d{6}Z)-table-(\d+)\.csv$")
    grouped: dict[str, list[Path]] = {}
    for path in raw_dir.glob("wikipedia-fr-2027-polls-*-table-*.csv"):
        match = pattern.match(path.name)
        if not match:
            continue
        grouped.setdefault(match.group(1), []).append(path)
    if not grouped:
        return []
    latest_key = sorted(grouped.keys())[-1]
    return sorted(grouped[latest_key], key=lambda path: int(pattern.match(path.name).group(2)))


def _parse_first_round_raw_wikipedia_table(table_path: Path, fallback_year: int) -> pd.DataFrame:
    frame = pd.read_csv(table_path, header=None, dtype=str, keep_default_na=False)
    if frame.shape[1] < 10 or str(frame.iat[0, 0]).strip() != "Sondeur":
        return pd.DataFrame()
    candidate_headers = [_strip_wikipedia_annotations(value) for value in frame.iloc[1, 3:].tolist()]
    rows: list[dict[str, object]] = []
    scenario_counters: dict[tuple[str, str], int] = {}
    table_match = re.search(r"table-(\d+)\.csv$", table_path.name)
    table_index = table_match.group(1) if table_match else "00"
    for row_index in range(4, len(frame)):
        pollster = frame.iat[row_index, 0]
        date_text = frame.iat[row_index, 1]
        sample_size = frame.iat[row_index, 2]
        if not _row_looks_like_poll(pollster, date_text, sample_size):
            continue
        fieldwork_start_date, fieldwork_end_date = _parse_fieldwork_dates(str(date_text), fallback_year=fallback_year)
        date_error = (
            None
            if fieldwork_start_date is not None or fieldwork_end_date is not None
            else "unparseable fieldwork date"
        )
        pollster_label = _normalize_company_name(str(pollster))
        key = (pollster_label, str(date_text))
        scenario_counters[key] = scenario_counters.get(key, 0) + 1
        scenario_index = scenario_counters[key]
        scenario_name = f"{pollster_label} · {date_text} · scénario {scenario_index}"
        poll_id = f"RAW-FR-{pollster_label.upper().replace(' ', '-')}-{table_index}-{row_index:03d}"

        for column_index, header_label in enumerate(candidate_headers, start=3):
            cell_text = str(frame.iat[row_index, column_index] or "").strip()
            estimate = _parse_raw_poll_percent(cell_text)
            is_not_tested = cell_text in {"", "-", "—"}
            candidate_fragment = re.sub(r"^\s*\d+(?:[.,]\d+)?\s*", "", _strip_wikipedia_annotations(cell_text)).strip()
            has_named_candidate = bool(re.search(r"[A-Za-zÀ-ÖØ-öø-ÿ]", candidate_fragment))
            candidate_label = candidate_fragment if has_named_candidate else header_label
            generic_header = header_label.startswith("Candidat ") or header_label in {"Autre", "Autres"}
            candidate_name, candidate_party, political_family = _extract_candidate_and_party_from_label(
                candidate_label,
                header_label if generic_header else None,
            )
            rows.append(
                {
                    "poll_id": poll_id,
                    "source_url": "https://fr.wikipedia.org/wiki/Liste_de_sondages_sur_l%27%C3%A9lection_pr%C3%A9sidentielle_fran%C3%A7aise_de_2027",
                    "source_name": "wikipedia_fr_raw_tables",
                    "polling_company": pollster_label,
                    "commissioner": None,
                    "media_partner": None,
                    "fieldwork_start_date": fieldwork_start_date,
                    "fieldwork_end_date": fieldwork_end_date,
                    "publication_date": fieldwork_end_date or fieldwork_start_date,
                    "sample_size": _parse_sample_size(sample_size),
                    "population": "unknown",
                    "collection_method": "unknown",
                    "quota_method": "unknown",
                    "round": "first_round",
                    "scenario_name": scenario_name,
                    "candidate_name": candidate_name,
                    "candidate_party": candidate_party,
                    "political_family": political_family,
                    "estimate_percent": estimate,
                    "lower_bound_percent": None,
                    "upper_bound_percent": None,
                    "margin_of_error": None,
                    "undecided_percent": None,
                    "abstention_estimate": None,
                    "registered_voters_basis": None,
                    "raw_text_context": cell_text,
                    "extraction_confidence": 0.55,
                    "fieldwork_date_raw": str(date_text),
                    "parse_status": (
                        "not_tested"
                        if is_not_tested
                        else "parsed"
                        if estimate is not None
                        else "unparsed_estimate"
                    ),
                    "parse_error": (
                        None
                        if is_not_tested
                        else "unparseable percentage"
                        if estimate is None
                        else date_error
                    ),
                }
            )
    return pd.DataFrame(rows)


def _parse_second_round_raw_wikipedia_table(table_path: Path, fallback_year: int) -> pd.DataFrame:
    frame = pd.read_csv(table_path, header=None, dtype=str, keep_default_na=False)
    if frame.shape[1] != 5 or str(frame.iat[0, 0]).strip() != "Sondeur":
        return pd.DataFrame()
    candidate_headers = [_strip_wikipedia_annotations(value) for value in frame.iloc[0, 3:5].tolist()]
    if any(str(header).startswith("Unnamed:") for header in candidate_headers):
        return pd.DataFrame()
    rows: list[dict[str, object]] = []
    table_match = re.search(r"table-(\d+)\.csv$", table_path.name)
    table_index = table_match.group(1) if table_match else "00"
    for row_index in range(3, len(frame)):
        pollster = frame.iat[row_index, 0]
        date_text = frame.iat[row_index, 1]
        sample_size = frame.iat[row_index, 2]
        if not _row_looks_like_poll(pollster, date_text, sample_size):
            continue
        fieldwork_start_date, fieldwork_end_date = _parse_fieldwork_dates(str(date_text), fallback_year=fallback_year)
        date_error = (
            None
            if fieldwork_start_date is not None or fieldwork_end_date is not None
            else "unparseable fieldwork date"
        )
        pollster_label = _normalize_company_name(str(pollster))
        scenario_name = f"{candidate_headers[0]} / {candidate_headers[1]}"
        poll_id = f"RAW-SR-{pollster_label.upper().replace(' ', '-')}-{table_index}-{row_index:03d}"
        for offset, candidate_label in enumerate(candidate_headers, start=3):
            cell_text = str(frame.iat[row_index, offset] or "").strip()
            estimate = _parse_raw_poll_percent(cell_text)
            is_not_tested = cell_text in {"", "-", "—"}
            candidate_name, candidate_party, political_family = _extract_candidate_and_party_from_label(candidate_label)
            rows.append(
                {
                    "poll_id": poll_id,
                    "source_url": "https://fr.wikipedia.org/wiki/Liste_de_sondages_sur_l%27%C3%A9lection_pr%C3%A9sidentielle_fran%C3%A7aise_de_2027",
                    "source_name": "wikipedia_fr_raw_tables",
                    "polling_company": pollster_label,
                    "commissioner": None,
                    "media_partner": None,
                    "fieldwork_start_date": fieldwork_start_date,
                    "fieldwork_end_date": fieldwork_end_date,
                    "publication_date": fieldwork_end_date or fieldwork_start_date,
                    "sample_size": _parse_sample_size(sample_size),
                    "population": "unknown",
                    "collection_method": "unknown",
                    "quota_method": "unknown",
                    "round": "second_round",
                    "scenario_name": scenario_name,
                    "candidate_name": candidate_name,
                    "candidate_party": candidate_party,
                    "political_family": political_family,
                    "estimate_percent": estimate,
                    "lower_bound_percent": None,
                    "upper_bound_percent": None,
                    "margin_of_error": None,
                    "undecided_percent": None,
                    "abstention_estimate": None,
                    "registered_voters_basis": None,
                    "raw_text_context": cell_text,
                    "extraction_confidence": 0.65,
                    "fieldwork_date_raw": str(date_text),
                    "parse_status": (
                        "not_tested"
                        if is_not_tested
                        else "parsed"
                        if estimate is not None
                        else "unparsed_estimate"
                    ),
                    "parse_error": (
                        None
                        if is_not_tested
                        else "unparseable percentage"
                        if estimate is None
                        else date_error
                    ),
                }
            )
    return pd.DataFrame(rows)


def raw_wikipedia_2027_tables_to_normalized_dataframe(raw_dir: Path) -> pd.DataFrame:
    table_files = _latest_wikipedia_fr_2027_table_files(raw_dir)
    if not table_files:
        return pd.DataFrame()
    timestamp_match = re.search(r"wikipedia-fr-2027-polls-(\d{4})", table_files[0].name)
    fallback_year = int(timestamp_match.group(1)) if timestamp_match else None
    frames: list[pd.DataFrame] = []
    for table_path in table_files:
        table_match = re.search(r"table-(\d+)\.csv$", table_path.name)
        table_index = int(table_match.group(1)) if table_match else 0
        first_round_years = {1: 2026, 2: 2026, 3: 2025, 4: 2024, 5: 2023, 6: 2023}
        table_fallback_year = first_round_years.get(table_index, fallback_year or 2026)
        parsed_first_round = _parse_first_round_raw_wikipedia_table(table_path, table_fallback_year)
        if not parsed_first_round.empty:
            frames.append(parsed_first_round)
            continue
        parsed_second_round = _parse_second_round_raw_wikipedia_table(table_path, fallback_year or 2026)
        if not parsed_second_round.empty:
            frames.append(parsed_second_round)
    if not frames:
        return pd.DataFrame()
    normalized = pd.concat(frames, ignore_index=True)
    normalized = _correct_poll_units_by_scenario(normalized)
    normalized = _rewrite_first_round_scenario_names(normalized)
    publication_dates = pd.to_datetime(normalized["publication_date"], errors="coerce")
    normalized["_publication_date"] = publication_dates
    recognized_future = normalized["_publication_date"].notna() & (
        normalized["_publication_date"].dt.date > date.today()
    )
    normalized = normalized.loc[~recognized_future].copy()
    normalized = normalized.drop(columns="_publication_date")
    ordered_columns = [
        "poll_id",
        "source_url",
        "source_name",
        "polling_company",
        "commissioner",
        "media_partner",
        "fieldwork_start_date",
        "fieldwork_end_date",
        "publication_date",
        "sample_size",
        "population",
        "collection_method",
        "quota_method",
        "round",
        "scenario_name",
        "candidate_name",
        "candidate_party",
        "political_family",
        "estimate_percent",
        "lower_bound_percent",
        "upper_bound_percent",
        "margin_of_error",
        "undecided_percent",
        "abstention_estimate",
        "registered_voters_basis",
        "raw_text_context",
        "extraction_confidence",
        *PARSING_DIAGNOSTIC_COLUMNS,
    ]
    return normalized.reindex(columns=ordered_columns)


def _candidate_order_from_sheet(frame: pd.DataFrame) -> list[str]:
    if frame.empty:
        return []
    candidate_column = "candidate" if "candidate" in frame.columns else frame.columns[1]
    return [_clean_candidate_name(candidate) for candidate in frame[candidate_column].dropna().tolist()]


def _scenario_name_from_v2_row(section: str, pollster: str, scenario_index: object) -> str:
    return f"{section} - scenario {scenario_index} - {pollster}"


SCENARIO_PRIORITY = [
    "Édouard Philippe",
    "Gabriel Attal",
    "Marine Le Pen",
    "Jordan Bardella",
    "Jean-Luc Mélenchon",
    "Raphaël Glucksmann",
    "Dominique de Villepin",
    "Bruno Retailleau",
    "François Ruffin",
    "Marine Tondelier",
    "François Bayrou",
    "Olivier Faure",
]


def _sort_candidates_for_label(candidates: set[str]) -> list[str]:
    return sorted(candidates, key=lambda name: (SCENARIO_PRIORITY.index(name) if name in SCENARIO_PRIORITY else 999, name))


def _humanize_second_round_name(name: str) -> str:
    parts = [part.strip() for part in re.split(r"\bvs\b|-", str(name), flags=re.IGNORECASE) if part.strip()]
    if len(parts) >= 2:
        left, _, _ = canonicalize_candidate_fields(parts[0])
        right, _, _ = canonicalize_candidate_fields(parts[1])
        return f"{left} / {right}"
    return str(name)


def _rewrite_first_round_scenario_names(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame

    working = frame.copy()
    first_round = working.loc[working["round"] == "first_round"].copy()
    if first_round.empty:
        return working

    bundle_keys = ["polling_company", "fieldwork_start_date", "fieldwork_end_date", "source_url", "round"]
    scenario_sets = (
        first_round.groupby(bundle_keys + ["poll_id", "scenario_name"], dropna=False)["candidate_name"]
        .apply(lambda series: set(str(value) for value in series.dropna().tolist() if not is_generic_bloc_label(value)))
        .reset_index(name="candidate_set")
    )

    renamed: dict[tuple[str, str], str] = {}
    for _, bundle in scenario_sets.groupby(bundle_keys, dropna=False):
        scenario_rows = bundle.to_dict(orient="records")
        scenario_sets_only = [row["candidate_set"] for row in scenario_rows if row["candidate_set"]]
        common_candidates = set.intersection(*scenario_sets_only) if scenario_sets_only else set()
        for row in scenario_rows:
            distinctive = row["candidate_set"] - common_candidates
            if distinctive:
                label_candidates = _sort_candidates_for_label(distinctive)
                label = "Hypothèse " + " / ".join(label_candidates[:3])
            else:
                label_candidates = _sort_candidates_for_label(row["candidate_set"])
                label = "Scénario " + " / ".join(label_candidates[:3]) if label_candidates else "Scénario premier tour"
            renamed[(row["poll_id"], row["scenario_name"])] = label

    working["scenario_name"] = working.apply(
        lambda row: renamed.get((row["poll_id"], row["scenario_name"]), row["scenario_name"])
        if row["round"] == "first_round"
        else _humanize_second_round_name(str(row["scenario_name"])),
        axis=1,
    )
    return working


def parse_first_round_sheet(frame: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for record in frame.to_dict(orient="records"):
        raw_summary = str(record.get("raw_context_or_structured_summary") or "")
        status = str(record.get("status") or "")
        if not raw_summary or "raw_block_needs_parser" in status:
            continue

        scenario_match = re.search(r"Scenario\s+([^:]+):", raw_summary, flags=re.IGNORECASE)
        scenario_name = scenario_match.group(1).strip() if scenario_match else f"Scenario {record.get('poll_id')}"
        after_colon = raw_summary.split(":", 1)[1] if ":" in raw_summary else raw_summary
        publication_date = _extract_publication_date_from_poll_id(str(record.get("poll_id") or ""))
        fieldwork_start_date, fieldwork_end_date = _parse_fieldwork_dates(
            str(record.get("fieldwork_dates") or ""),
            fallback_year=_publication_year_from_poll_id(record.get("poll_id")),
        )

        for chunk in after_colon.split(";"):
            piece = chunk.strip()
            if not piece or "—" in piece:
                continue
            candidate_match = re.match(r"(.+?)\s+(-?\d+(?:[.,]\d+)?)$", piece)
            if not candidate_match:
                continue
            raw_candidate, value = candidate_match.groups()
            candidate_party, political_family, candidate_name = _guess_party_and_family(raw_candidate)
            rows.append(
                {
                    "poll_id": record.get("poll_id"),
                    "source_url": record.get("source_url"),
                    "source_name": "wikipedia_fr_excel_extraction",
                    "polling_company": _normalize_company_name(str(record.get("polling_company") or "")),
                    "commissioner": None,
                    "media_partner": None,
                    "fieldwork_start_date": fieldwork_start_date,
                    "fieldwork_end_date": fieldwork_end_date,
                    "publication_date": publication_date,
                    "sample_size": record.get("sample_size"),
                    "population": "unknown",
                    "collection_method": "unknown",
                    "quota_method": "unknown",
                    "round": "first_round",
                    "scenario_name": scenario_name,
                    "candidate_name": candidate_name,
                    "candidate_party": candidate_party,
                    "political_family": political_family,
                    "estimate_percent": str(value).replace(",", "."),
                    "lower_bound_percent": None,
                    "upper_bound_percent": None,
                    "margin_of_error": None,
                    "undecided_percent": None,
                    "abstention_estimate": None,
                    "registered_voters_basis": None,
                    "raw_text_context": raw_summary,
                    "extraction_confidence": 0.75,
                }
            )
    return pd.DataFrame(rows)


def parse_second_round_sheet(frame: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for record in frame.to_dict(orient="records"):
        publication_date = _extract_publication_date_from_poll_id(str(record.get("poll_id") or ""))
        fieldwork_start_date, fieldwork_end_date = _parse_fieldwork_dates(
            str(record.get("fieldwork_dates") or ""),
            fallback_year=_publication_year_from_poll_id(record.get("poll_id")),
        )
        scenario_name = str(record.get("hypothesis") or f"Second round {record.get('poll_id')}")
        candidates = [
            (record.get("candidate_a"), record.get("candidate_a_percent")),
            (record.get("candidate_b"), record.get("candidate_b_percent")),
        ]
        for candidate_label, estimate in candidates:
            _, political_family, candidate_name = _guess_party_and_family(str(candidate_label))
            rows.append(
                {
                    "poll_id": record.get("poll_id"),
                    "source_url": record.get("source_url"),
                    "source_name": "wikipedia_fr_excel_extraction",
                    "polling_company": _normalize_company_name(str(record.get("polling_company") or "")),
                    "commissioner": None,
                    "media_partner": None,
                    "fieldwork_start_date": fieldwork_start_date,
                    "fieldwork_end_date": fieldwork_end_date,
                    "publication_date": publication_date,
                    "sample_size": record.get("sample_size"),
                    "population": "unknown",
                    "collection_method": "unknown",
                    "quota_method": "unknown",
                    "round": "second_round",
                    "scenario_name": scenario_name,
                    "candidate_name": candidate_name,
                    "candidate_party": None,
                    "political_family": political_family,
                    "estimate_percent": estimate,
                    "lower_bound_percent": None,
                    "upper_bound_percent": None,
                    "margin_of_error": None,
                    "undecided_percent": None,
                    "abstention_estimate": None,
                    "registered_voters_basis": None,
                    "raw_text_context": f"{candidate_name} vs scenario {scenario_name}",
                    "extraction_confidence": 0.9,
                }
            )
    return pd.DataFrame(rows)


def parse_second_round_structured_sheet(frame: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for index, record in enumerate(frame.to_dict(orient="records"), start=1):
        fieldwork_start_date, fieldwork_end_date = _parse_fieldwork_dates(
            str(record.get("fieldwork") or ""),
            fallback_year=_publication_year_from_poll_id(record.get("poll_id")),
        )
        scenario_name = str(record.get("scenario") or f"second_round_{index}")
        poll_id = f"V2-SR-{_normalize_company_name(str(record.get('pollster') or 'unknown')).upper().replace(' ', '-')}-{index:03d}"
        for prefix in ("a", "b"):
            candidate_label = _clean_candidate_name(str(record.get(f"candidate_{prefix}") or ""))
            candidate_party = record.get(f"candidate_{prefix}_party")
            _, political_family, candidate_name = _guess_party_and_family(candidate_label)
            candidate_name, candidate_party, political_family = canonicalize_candidate_fields(
                candidate_name, candidate_party, political_family
            )
            rows.append(
                {
                    "poll_id": poll_id,
                    "source_url": record.get("source_url"),
                    "source_name": "wikipedia_excel_v2",
                    "polling_company": _normalize_company_name(str(record.get("pollster") or "")),
                    "commissioner": None,
                    "media_partner": None,
                    "fieldwork_start_date": fieldwork_start_date,
                    "fieldwork_end_date": fieldwork_end_date,
                    "publication_date": fieldwork_end_date or fieldwork_start_date,
                    "sample_size": _parse_sample_size(record.get("sample_size")),
                    "population": "unknown",
                    "collection_method": "unknown",
                    "quota_method": "unknown",
                    "round": "second_round",
                    "scenario_name": scenario_name,
                    "candidate_name": candidate_name,
                    "candidate_party": candidate_party,
                    "political_family": political_family,
                    "estimate_percent": record.get(f"candidate_{prefix}_score"),
                    "lower_bound_percent": None,
                    "upper_bound_percent": None,
                    "margin_of_error": None,
                    "undecided_percent": None,
                    "abstention_estimate": None,
                    "registered_voters_basis": None,
                    "raw_text_context": f"{scenario_name} / {candidate_name}",
                    "extraction_confidence": 0.9,
                }
            )
    return pd.DataFrame(rows)


def parse_first_round_raw_vectors_sheet(
    frame: pd.DataFrame,
    candidate_order_2025plus: list[str],
    candidate_order_until_2025: list[str],
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for index, record in enumerate(frame.to_dict(orient="records"), start=1):
        section = str(record.get("source_section") or "first_round")
        order = candidate_order_2025plus if "2025" in section or "March 2025 onwards" in section else candidate_order_until_2025
        tokens = _tokenize_score_vector(str(record.get("scores_raw_vector") or ""))
        fieldwork_start_date, fieldwork_end_date = _parse_fieldwork_dates(
            str(record.get("fieldwork") or ""),
            fallback_year=_publication_year_from_poll_id(record.get("poll_id")),
        )
        poll_id = f"V2-FR-{_normalize_company_name(str(record.get('pollster') or 'unknown')).upper().replace(' ', '-')}-{index:03d}"
        scenario_name = _scenario_name_from_v2_row(section, str(record.get("pollster") or ""), record.get("scenario_index"))
        vector_mismatch = len(order) != len(tokens)
        for position, (candidate_name, token) in enumerate(
            zip_longest(order, tokens),
            start=1,
        ):
            if token is None:
                continue
            candidate_name = candidate_name or f"Valeur non attribuée {position}"
            estimate = None if token == "-" else token.replace(",", ".")
            parse_status = (
                "not_tested"
                if token == "-"
                else "vector_length_mismatch"
                if vector_mismatch
                else "parsed"
            )
            candidate_party, political_family, normalized_candidate = _guess_party_and_family(_clean_candidate_name(candidate_name))
            normalized_candidate, candidate_party, political_family = canonicalize_candidate_fields(
                normalized_candidate, candidate_party, political_family
            )
            rows.append(
                {
                    "poll_id": poll_id,
                    "source_url": record.get("source_url"),
                    "source_name": "wikipedia_excel_v2",
                    "polling_company": _normalize_company_name(str(record.get("pollster") or "")),
                    "commissioner": None,
                    "media_partner": None,
                    "fieldwork_start_date": fieldwork_start_date,
                    "fieldwork_end_date": fieldwork_end_date,
                    "publication_date": fieldwork_end_date or fieldwork_start_date,
                    "sample_size": _parse_sample_size(record.get("sample_size")),
                    "population": "unknown",
                    "collection_method": "unknown",
                    "quota_method": "unknown",
                    "round": "first_round",
                    "scenario_name": scenario_name,
                    "candidate_name": normalized_candidate,
                    "candidate_party": candidate_party,
                    "political_family": political_family,
                    "estimate_percent": estimate,
                    "lower_bound_percent": None,
                    "upper_bound_percent": None,
                    "margin_of_error": None,
                    "undecided_percent": None,
                    "abstention_estimate": None,
                    "registered_voters_basis": None,
                    "raw_text_context": str(record.get("scores_raw_vector") or ""),
                    "extraction_confidence": 0.8,
                    "parse_status": parse_status,
                    "parse_error": (
                        f"candidate_count={len(order)} token_count={len(tokens)}"
                        if vector_mismatch
                        else None
                    ),
                }
            )
    return pd.DataFrame(rows)


def parse_scenario_polling_raw_sheet(frame: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for index, record in enumerate(frame.to_dict(orient="records"), start=1):
        order_reference = str(record.get("candidate_order_reference") or "")
        if "Generic" in order_reference:
            order = [chunk.strip() for chunk in order_reference.replace("Generic", "").split("/") if chunk.strip()]
        elif "LePen_runs_candidate_order" in order_reference:
            order = [
                "Arthaud", "Roussel", "Mélenchon", "Glucksmann", "Tondelier", "Attal",
                "Philippe", "Retailleau", "Dupont-Aignan", "Le Pen", "Zemmour"
            ]
        else:
            order = []
        tokens = _tokenize_score_vector(str(record.get("scores_raw_vector") or ""))
        fieldwork_start_date, fieldwork_end_date = _parse_fieldwork_dates(
            str(record.get("fieldwork") or ""),
            fallback_year=_publication_year_from_poll_id(record.get("poll_id")),
        )
        poll_id = f"V2-SP-{_normalize_company_name(str(record.get('pollster') or 'unknown')).upper().replace(' ', '-')}-{index:03d}"
        scenario_name = _scenario_name_from_v2_row(str(record.get("section") or "scenario"), str(record.get("pollster") or ""), record.get("scenario_index"))
        vector_mismatch = len(order) != len(tokens)
        for position, (candidate_name, token) in enumerate(
            zip_longest(order, tokens),
            start=1,
        ):
            if token is None:
                continue
            candidate_name = candidate_name or f"Valeur non attribuée {position}"
            estimate = None if token == "-" else token.replace(",", ".")
            parse_status = (
                "not_tested"
                if token == "-"
                else "vector_length_mismatch"
                if vector_mismatch
                else "parsed"
            )
            candidate_party, political_family, normalized_candidate = _guess_party_and_family(_clean_candidate_name(candidate_name))
            normalized_candidate, candidate_party, political_family = canonicalize_candidate_fields(
                normalized_candidate, candidate_party, political_family
            )
            rows.append(
                {
                    "poll_id": poll_id,
                    "source_url": record.get("source_url"),
                    "source_name": "wikipedia_excel_v2",
                    "polling_company": _normalize_company_name(str(record.get("pollster") or "")),
                    "commissioner": None,
                    "media_partner": None,
                    "fieldwork_start_date": fieldwork_start_date,
                    "fieldwork_end_date": fieldwork_end_date,
                    "publication_date": fieldwork_end_date or fieldwork_start_date,
                    "sample_size": _parse_sample_size(record.get("sample_size")),
                    "population": "unknown",
                    "collection_method": "unknown",
                    "quota_method": "unknown",
                    "round": "first_round",
                    "scenario_name": scenario_name,
                    "candidate_name": normalized_candidate,
                    "candidate_party": candidate_party,
                    "political_family": political_family,
                    "estimate_percent": estimate,
                    "lower_bound_percent": None,
                    "upper_bound_percent": None,
                    "margin_of_error": None,
                    "undecided_percent": None,
                    "abstention_estimate": None,
                    "registered_voters_basis": None,
                    "raw_text_context": str(record.get("scores_raw_vector") or ""),
                    "extraction_confidence": 0.7,
                    "parse_status": parse_status,
                    "parse_error": (
                        f"candidate_count={len(order)} token_count={len(tokens)}"
                        if vector_mismatch
                        else None
                    ),
                }
            )
    return pd.DataFrame(rows)


def workbook_to_normalized_dataframe(workbook_path: Path) -> pd.DataFrame:
    sheets = load_workbook_sheets(workbook_path)
    frames: list[pd.DataFrame] = []
    if "first_round_raw_vectors" in sheets or "second_round_structured" in sheets:
        candidate_order_2025plus = _candidate_order_from_sheet(sheets.get("candidate_order_2025plus", pd.DataFrame()))
        candidate_order_until_2025 = _candidate_order_from_sheet(sheets.get("candidate_order_until_2025", pd.DataFrame()))
        if "first_round_raw_vectors" in sheets:
            frames.append(
                parse_first_round_raw_vectors_sheet(
                    sheets["first_round_raw_vectors"],
                    candidate_order_2025plus=candidate_order_2025plus,
                    candidate_order_until_2025=candidate_order_until_2025,
                )
            )
        if "second_round_structured" in sheets:
            frames.append(parse_second_round_structured_sheet(sheets["second_round_structured"]))
        if "scenario_polling_raw" in sheets:
            frames.append(parse_scenario_polling_raw_sheet(sheets["scenario_polling_raw"]))
    else:
        if "first_round" in sheets:
            frames.append(parse_first_round_sheet(sheets["first_round"]))
        if "second_round" in sheets:
            frames.append(parse_second_round_sheet(sheets["second_round"]))
    if not frames:
        return pd.DataFrame()
    normalized = pd.concat(frames, ignore_index=True)
    normalized = _correct_poll_units_by_scenario(normalized)
    normalized = _rewrite_first_round_scenario_names(normalized)
    ordered_columns = [
        "poll_id",
        "source_url",
        "source_name",
        "polling_company",
        "commissioner",
        "media_partner",
        "fieldwork_start_date",
        "fieldwork_end_date",
        "publication_date",
        "sample_size",
        "population",
        "collection_method",
        "quota_method",
        "round",
        "scenario_name",
        "candidate_name",
        "candidate_party",
        "political_family",
        "estimate_percent",
        "lower_bound_percent",
        "upper_bound_percent",
        "margin_of_error",
        "undecided_percent",
        "abstention_estimate",
        "registered_voters_basis",
        "raw_text_context",
        "extraction_confidence",
        *PARSING_DIAGNOSTIC_COLUMNS,
    ]
    return normalized.reindex(columns=ordered_columns)
